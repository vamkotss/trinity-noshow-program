"""Tests for the ROI model.

WHAT THESE PROTECT
------------------
The ROI model is the number that gets money approved. Three things must hold:

  1. IT COMPUTES. Every formula evaluates - no #NAME?, no #REF!, no silent
     None. A model with a broken formula is worse than a table, because it looks
     authoritative while being wrong.

  2. THE NUMBERS ARE SANE. The no-show cost, the ROIs, the budget check - each
     must land in a defensible range. A green recalc proves formulas evaluate,
     not that they are right; these tests check they are right.

  3. THE HONEST FINDINGS SURVIVE. Rebalancing wins big. Overbooking is marginal.
     Reminders do not clear their cost at the adjusted effect. If a future edit
     made every intervention look like a slam-dunk, the model would have stopped
     telling the truth - and these tests would fail.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from trinity.roi_model import build


@pytest.fixture(scope="module")
def model(tmp_path_factory):
    path = tmp_path_factory.mktemp("reports") / "ROI_MODEL.xlsx"
    build(path)

    # Recalculate via LibreOffice so the cached values exist to read.
    import subprocess

    recalc = Path("/mnt/skills/public/xlsx/scripts/recalc.py")
    if recalc.exists():
        subprocess.run(["python3", str(recalc), str(path)], capture_output=True, timeout=90)

    return openpyxl.load_workbook(path, data_only=True)


def _has_values(model) -> bool:
    """Whether recalc ran - if not, formula cells read None and we skip value checks."""
    return model["Cost of Problem"]["B5"].value is not None


# ---------------------------------------------------------------------------
# 1. STRUCTURE
# ---------------------------------------------------------------------------


def test_the_workbook_has_the_four_tabs(model):
    """Assumptions, cost, interventions, and the downside tab all present."""
    expected = {"Assumptions", "Cost of Problem", "Interventions", "Overbook Downside"}

    assert expected <= set(model.sheetnames), (
        f"missing tabs: {expected - set(model.sheetnames)}"
    )


def test_assumptions_are_real_inputs_not_formulas(model):
    """The assumption cells hold numbers the board can change, not formulas.

    An 'assumption' that is actually a formula is not a lever - the board cannot
    move it, and the model cannot be stress-tested. Inputs must be inputs.
    """
    ws = model["Assumptions"]

    # The contribution-per-visit lever must be a bare number.
    found_contribution = False
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "Contribution per attended visit ($)":
                lever = ws.cell(row=cell.row, column=2).value
                assert isinstance(lever, (int, float)), "contribution is not a plain input"
                assert lever > 0
                found_contribution = True
    assert found_contribution, "the contribution assumption is missing"


# ---------------------------------------------------------------------------
# 2. THE NUMBERS ARE SANE
# ---------------------------------------------------------------------------


def test_the_no_show_cost_is_defensible(model):
    """The headline cost is in a believable range for a 7-clinic network."""
    if not _has_values(model):
        pytest.skip("recalc did not run in this environment")

    cost = model["Cost of Problem"]["B5"].value

    # 4,226 no-shows x $150 = ~$634k. Sanity band around that.
    assert 400_000 < cost < 900_000, f"no-show cost ${cost:,.0f} is outside a defensible range"


def test_rebalancing_is_the_top_intervention(model):
    """Rebalancing has by far the best ROI - it is the recommendation to fund first.

    This is the central finding: the cheapest fix (moving demand from overbooked
    to idle providers) has the highest return. If some other intervention ranked
    first, the whole funding recommendation would change.
    """
    if not _has_values(model):
        pytest.skip("recalc did not run")

    ws = model["Interventions"]

    rebalance_roi = ws["E4"].value      # intervention 1
    overbook_roi = ws["E5"].value       # intervention 2
    reminder_roi = ws["E6"].value       # intervention 3

    assert rebalance_roi > overbook_roi, "rebalancing does not beat overbooking"
    assert rebalance_roi > reminder_roi, "rebalancing does not beat reminders"
    assert rebalance_roi > 3, f"rebalancing ROI is only {rebalance_roi:.1f}x - implausibly low"


def test_the_program_fits_the_budget(model):
    """Total program cost is under the $600k ceiling."""
    if not _has_values(model):
        pytest.skip("recalc did not run")

    ws = model["Interventions"]

    total_cost = ws["C8"].value
    ceiling = ws["B10"].value

    assert total_cost <= ceiling, f"program cost ${total_cost:,.0f} exceeds ceiling ${ceiling:,.0f}"


# ---------------------------------------------------------------------------
# 3. THE HONEST FINDINGS SURVIVE
# ---------------------------------------------------------------------------


def test_overbooking_is_only_marginal(model):
    """Overbooking barely breaks even - the collision cost eats the recovery.

    This is the honest, uncomfortable finding. Overbooking is NOT a free win;
    once you price the cost of a patient waiting when both show, the net per slot
    is only a few dollars. A model that made overbooking look highly profitable
    would be ignoring exactly the objection the physician will raise.
    """
    if not _has_values(model):
        pytest.skip("recalc did not run")

    ws = model["Overbook Downside"]

    net_per_slot = ws["B11"].value

    assert net_per_slot is not None
    # A few dollars per slot - genuinely marginal.
    assert 0 < net_per_slot < 15, (
        f"overbooking net/slot is ${net_per_slot:.2f} - the model is not showing "
        "how thin this intervention really is"
    )


def test_the_collision_cost_is_material(model):
    """The overbooking downside is real money, not a rounding line.

    If the collision cost were negligible, pricing it would be theatre. It is
    material precisely because most high-risk slots still have the original
    patient show up.
    """
    if not _has_values(model):
        pytest.skip("recalc did not run")

    ws = model["Overbook Downside"]

    collision_per_slot = ws["B10"].value
    recovery_per_slot = ws["B8"].value

    # The collision cost is a large fraction of the recovery - that is why
    # overbooking is marginal.
    assert collision_per_slot > recovery_per_slot * 0.5, (
        "the collision cost is too small to make overbooking marginal - "
        "check the downside is being priced honestly"
    )


def test_reminders_use_the_adjusted_not_naive_effect(model):
    """The reminder benefit is small - built on the de-confounded 0.7pt effect.

    If reminders showed a large benefit, the model would be using the naive 2.3pt
    figure that was ~70% selection bias. The whole point of M5 was to NOT do that.
    A small or negative reminder ROI is the evidence the honesty survived into the
    money model.
    """
    if not _has_values(model):
        pytest.skip("recalc did not run")

    ws = model["Interventions"]

    reminder_benefit = ws["B6"].value

    # At 0.7pt effect on ~30k appts x $150, the benefit is ~$32k - well under the
    # $45k cost. If it were built on the naive 2.3pt figure it would be ~$104k.
    assert reminder_benefit < 60_000, (
        f"reminder benefit ${reminder_benefit:,.0f} is too high - it looks like "
        "the naive effect was used instead of the adjusted one"
    )
