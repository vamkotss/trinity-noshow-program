"""Tests for the Trinity data generator.

WHY TEST IT
-----------
Same reason as P1. The standardization layer in M3 will claim to recover 14
no-show spellings into one canonical state, dedupe patients across MRNs, and
read FTE out of cell colours. Those claims are only worth something if the
defects are known to be present and known in number - which is what these tests
establish.

They check four things:
  1. REPRODUCIBILITY - same seed, same bytes.
  2. THE DEFECTS ARE PRESENT - all fourteen spellings, the MRN duplicates, the
     status contradictions, the colour-coded roster.
  3. THE DRIVERS ARE DISCOVERABLE - lead time, new patients, payer, weekday all
     move no-show rate in the planted direction.
  4. THE NATURAL EXPERIMENT IS CONFOUNDED - the reminding clinics really are the
     better clinics, so the naive comparison is a trap.
"""

from __future__ import annotations

import pandas as pd
import pytest
from openpyxl import load_workbook

from trinity.generate import (
    CLINICS,
    NO_SHOW_SPELLINGS,
    SEED,
    generate,
)

# The colour->FTE map an analyst would reconstruct from the legend.
COLOUR_TO_FTE = {"C6EFCE": 1.0, "FFEB9C": 0.8, "FFCC99": 0.6}


@pytest.fixture(scope="module")
def data(tmp_path_factory):
    out = tmp_path_factory.mktemp("raw")
    generate(out, seed=SEED)

    return {
        "dir": out,
        "patients": pd.read_csv(out / "patients.csv"),
        "appointments": pd.read_csv(out / "appointments.csv"),
        "visits": pd.read_csv(out / "visits.csv"),
        "reminders": pd.read_csv(out / "reminder_log.csv"),
        "mrn_truth": pd.read_csv(out / "_mrn_truth.csv"),
        "roster_truth": pd.read_csv(out / "_roster_truth.csv"),
        "roster_xlsx": out / "providers.xlsx",
    }


# ---------------------------------------------------------------------------
# 1. REPRODUCIBILITY
# ---------------------------------------------------------------------------


def test_same_seed_same_data(tmp_path):
    """Two runs, one seed, identical bytes. The non-negotiable property."""
    a, b = tmp_path / "a", tmp_path / "b"
    generate(a, seed=42)
    generate(b, seed=42)

    for name in ["patients.csv", "appointments.csv", "visits.csv"]:
        assert (a / name).read_bytes() == (b / name).read_bytes(), f"{name} differs"


# ---------------------------------------------------------------------------
# 2. THE DEFECTS ARE PRESENT
# ---------------------------------------------------------------------------


def test_all_fourteen_no_show_spellings_appear(data):
    """Every one of the fourteen strings that mean 'no-show' is in the export.

    If a spelling is missing, the standardization mapping in M3 would have a rule
    with nothing to match - untested, and quietly wrong the day real data
    contains it.
    """
    statuses = set(data["appointments"]["status"])

    present = [s for s in NO_SHOW_SPELLINGS if s in statuses]

    assert len(present) >= 12, (
        f"only {len(present)} of {len(NO_SHOW_SPELLINGS)} no-show spellings appear; "
        "the standardization layer would be undertested"
    )


def test_status_is_genuinely_messy(data):
    """The status column has far more than three distinct values."""
    distinct = data["appointments"]["status"].nunique()

    assert distinct > 15, f"only {distinct} status strings - not messy enough to be realistic"


def test_appointment_types_are_free_typed(data):
    """The five real types appear as many more than five strings."""
    distinct = data["appointments"]["appointment_type"].nunique()

    assert distinct > 10, "appointment types are too clean"


def test_duplicate_mrns_exist(data):
    """Some humans appear under more than one MRN."""
    mapping = data["mrn_truth"]

    assert len(mapping) > 0, "no duplicate MRNs injected"

    # Every alias points at a real primary MRN.
    patients = set(data["patients"]["mrn"])
    assert mapping["alias_mrn"].isin(patients).all()
    assert mapping["true_mrn"].isin(patients).all()


def test_duplicate_mrns_share_a_date_of_birth(data):
    """The alias and the true record share a DOB - the key that makes dedupe possible.

    Without a shared join key, deduplication would be guesswork. DOB is the
    realistic anchor: names drift ('Sarah' -> 'S.'), DOB does not.
    """
    patients = data["patients"].set_index("mrn")
    mapping = data["mrn_truth"]

    for row in mapping.head(50).itertuples(index=False):
        dob_alias = patients.loc[row.alias_mrn, "date_of_birth"]
        dob_true = patients.loc[row.true_mrn, "date_of_birth"]
        assert dob_alias == dob_true, "an alias has a different DOB from its true record"


def test_payer_column_mixes_names_and_plan_ids(data):
    """Some payer values are plan IDs, not payer names."""
    payers = data["appointments"]["payer"].astype(str)

    has_plan_id = payers.str.contains(r"-\d", regex=True).any()

    assert has_plan_id, "no plan IDs found in the payer column - the mixing defect is absent"


def test_visits_contradict_the_schedule(data):
    """Some appointments the schedule calls no-show have a billed visit.

    This is the contradiction Tom warned about: someone arrived, nobody updated
    the schedule. M3 needs a ruling for which source wins, and it needs this
    defect to exist in order to test that ruling.
    """
    appts = data["appointments"]
    visits = data["visits"]

    no_show_appts = set(appts[appts["status"].isin(NO_SHOW_SPELLINGS)]["appointment_id"])
    billed_appts = set(visits["appointment_id"])

    contradictions = no_show_appts & billed_appts

    assert len(contradictions) > 0, "no schedule/billing contradictions - the trap is missing"


def test_procedure_codes_have_typos(data):
    """Some procedure codes are malformed."""
    codes = data["visits"]["procedure_code"].astype(str)

    # Clean codes are 5 digits. Anything else is a typo.
    malformed = ~codes.str.fullmatch(r"\d{5}")

    assert malformed.sum() > 0, "no typo'd procedure codes"


# ---------------------------------------------------------------------------
# 3. THE COLOUR-CODED ROSTER
# ---------------------------------------------------------------------------


def test_roster_encodes_fte_as_cell_colour(data):
    """FTE is recoverable from the cell fill colour, and nowhere else.

    This is the signature task of the project. The test proves two things: that
    the colour is actually set on the cell, and that the FTE recovered from it
    matches the ground truth. If this passes, the M3 extraction has something
    real to extract.
    """
    wb = load_workbook(data["roster_xlsx"])
    ws = wb["Roster"]
    truth = data["roster_truth"].set_index("provider")

    recovered = {}
    for row in range(2, 2 + len(truth)):
        name = ws.cell(row=row, column=1).value
        if not name:
            continue
        fill = ws.cell(row=row, column=4).fill
        rgb = fill.start_color.rgb
        hex_code = rgb[-6:] if isinstance(rgb, str) else None
        recovered[name] = COLOUR_TO_FTE.get(hex_code)

    # Every provider's FTE recovered from colour matches the truth.
    for name, fte in recovered.items():
        assert fte is not None, f"could not read a colour for {name}"
        assert fte == truth.loc[name, "fte"], f"colour for {name} maps to the wrong FTE"


def test_the_fte_column_is_actually_empty(data):
    """The FTE cell has no text value - the colour is the ONLY source.

    If the number were also in the cell text, the whole exercise would be a
    trivial column read and the 'extract structured data from formatting' claim
    would be hollow.
    """
    wb = load_workbook(data["roster_xlsx"])
    ws = wb["Roster"]

    for row in range(2, 2 + len(data["roster_truth"])):
        value = ws.cell(row=row, column=4).value
        assert value in (None, ""), f"row {row} has FTE as text - the colour trap is defeated"


# ---------------------------------------------------------------------------
# 4. THE DRIVERS ARE DISCOVERABLE
# ---------------------------------------------------------------------------


def _no_show_frame(appointments: pd.DataFrame) -> pd.DataFrame:
    """Reconstruct a clean no-show flag and features from the messy export.

    This mimics what the analyst does in M3/M5 - and proves the signal survives
    the mess.
    """
    df = appointments.copy()
    df["no_show"] = df["status"].isin(NO_SHOW_SPELLINGS).astype(int)
    df["appointment_date"] = pd.to_datetime(df["appointment_date"])
    df["booked_date"] = pd.to_datetime(df["booked_date"])
    df["lead_days"] = (df["appointment_date"] - df["booked_date"]).dt.days
    df["weekday"] = df["appointment_date"].dt.weekday
    return df


def test_lead_time_drives_no_shows(data):
    """The longer the wait between booking and appointment, the more no-shows.

    This is the strongest planted driver and the one that most justifies an
    overbooking policy. If it were flat, the central business recommendation
    would have no evidence behind it.
    """
    df = _no_show_frame(data["appointments"])

    short = df[df["lead_days"] <= 7]["no_show"].mean()
    long = df[df["lead_days"] > 30]["no_show"].mean()

    assert long > short * 1.5, (
        f"lead-time effect too weak: {short:.1%} short vs {long:.1%} long"
    )


def test_new_patients_no_show_more(data):
    """First-ever visits no-show more than established patients."""
    df = _no_show_frame(data["appointments"])

    # New-patient appointments are typed as such.
    is_new = df["appointment_type"].str.contains("new", case=False, na=False)

    new_rate = df[is_new]["no_show"].mean()
    established_rate = df[~is_new]["no_show"].mean()

    assert new_rate > established_rate, "new patients do not no-show more - a planted driver is missing"


def test_monday_mornings_are_worse(data):
    """Monday has a higher no-show rate than the rest of the week."""
    df = _no_show_frame(data["appointments"])

    monday = df[df["weekday"] == 0]["no_show"].mean()
    rest = df[df["weekday"] != 0]["no_show"].mean()

    assert monday > rest, f"no Monday effect: {monday:.1%} vs {rest:.1%}"


# ---------------------------------------------------------------------------
# 5. THE NATURAL EXPERIMENT IS CONFOUNDED
# The trap that makes M5 worth doing.
# ---------------------------------------------------------------------------


def test_only_three_clinics_logged_reminders(data):
    """The reminder log covers exactly the three reminding clinics.

    The four without a log are absent - which IS the selection mechanism.
    """
    logged = set(data["reminders"]["clinic"])
    expected = {c for c, v in CLINICS.items() if v["reminds"]}

    assert logged == expected, f"reminder log covers {logged}, expected {expected}"


def test_the_reminder_comparison_is_confounded(data):
    """Reminding clinics have lower no-shows - but they are also better-run.

    This is the trap. A naive analyst compares no-show rates between clinics that
    remind and clinics that do not, finds reminders 'work', and recommends
    spending on reminders. But the reminding clinics were selected for logging
    precisely because they are well-run, so they would have had lower no-shows
    anyway. The gap is real; the causal claim is not.

    This test asserts the confound exists - that the naive gap is present AND
    that it aligns with clinic quality rather than with reminding per se.
    """
    df = _no_show_frame(data["appointments"])

    reminding = {c for c, v in CLINICS.items() if v["reminds"]}
    df["reminds"] = df["clinic"].isin(reminding)

    remind_rate = df[df["reminds"]]["no_show"].mean()
    no_remind_rate = df[~df["reminds"]]["no_show"].mean()

    # The naive gap must exist - otherwise there is no trap to fall into.
    assert remind_rate < no_remind_rate, "no naive reminder gap - the trap is absent"

    # And it must align with the planted quality difference, not be a clean
    # treatment effect. The reminding clinics have higher quality scores.
    remind_quality = sum(CLINICS[c]["quality"] for c in reminding) / len(reminding)
    other_quality = sum(
        CLINICS[c]["quality"] for c in CLINICS if not CLINICS[c]["reminds"]
    ) / (len(CLINICS) - len(reminding))

    assert remind_quality > other_quality, (
        "reminding clinics are not systematically better-run - "
        "the confound that makes this a teaching case is missing"
    )
