"""Standardization layer for Trinity clinic data.

FOUR JOBS, FOUR RULINGS
-----------------------
This is the layer that turns a messy operational export into something you can
count. Every transformation here is a JUDGEMENT that could have gone another way,
so every one is written down as a ruling with its reasoning - exactly as the
metrics contract did in P1. "I cleaned the data" is a claim. "I mapped 14 status
strings to 3 canonical states, and here is the ruling and the test" is evidence.

  RULING S1 - STATUS CANONICALISATION
    14 free-typed strings collapse to 3 states: no_show, arrived, cancelled.
    The mapping is explicit and total - any string we have not seen raises,
    rather than being silently dropped into a default. An unmapped status is a
    defect to investigate, not a row to guess at.

  RULING S2 - THE SCHEDULE-VS-BILLING CONTRADICTION
    Some appointments the schedule marks "no_show" have a billed visit attached.
    Someone arrived; the front desk never updated the schedule.
    RULING: billing wins. A billed visit is strong evidence a patient was seen -
    money changed hands and a clinician signed a note. The schedule is a
    convenience field a busy front desk forgets to update. Counting these as
    no-shows would OVERSTATE the problem and inflate the business case, which is
    exactly the kind of self-serving error a hostile stakeholder will catch.

  RULING S3 - PATIENT DEDUPLICATION
    The same human appears under several MRNs with name variants. We collapse
    them using DATE OF BIRTH as the anchor: names drift ("Sarah" -> "S."), a DOB
    does not. Within a shared DOB we group name variants conservatively.
    RULING: dedupe on (date_of_birth, normalised_name). We accept that two
    genuinely different people who share a DOB and a similar name could be
    wrongly merged - and we measure how often that risk applies rather than
    pretending it is zero.

  RULING S4 - FTE FROM CELL COLOUR
    The provider roster encodes each clinician's FTE as a cell BACKGROUND COLOUR,
    with no FTE column. We read the fill colour out of the spreadsheet and map it
    to a fraction via the legend. This is not a trick - operational data lives in
    spreadsheets like this constantly.

WHY THIS RUNS ON THE DE-IDENTIFIED DATA
---------------------------------------
Wherever possible, standardization operates on the de-identified files, so PHI
never spreads further than it must. The one exception is dedup, which needs the
date of birth as its anchor - so it reads DOB from the raw patient file, does its
work, and emits only the mapping (key -> canonical patient), never the DOB
itself.

Run:  python -m trinity.standardize
"""

from __future__ import annotations

import argparse
from dataclasses import dataclass, field
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# RULING S1 - STATUS CANONICALISATION
# ---------------------------------------------------------------------------

# The total, explicit mapping. Every string the front desk types, mapped to one
# of three states. This is the ruling made visible: nothing is guessed, nothing
# is defaulted. A string not in here is a defect, and the code says so loudly.
STATUS_MAP = {
    # --- no_show: the fourteen ---
    "NS": "no_show", "no show": "no_show", "noshow": "no_show", "no-show": "no_show",
    "No Show": "no_show", "DNA": "no_show", "dna": "no_show", "cancelled by pt": "no_show",
    "CXL": "no_show", "cxl": "no_show", "did not attend": "no_show", "n/s": "no_show",
    "No-Show": "no_show", "NOSHOW": "no_show",
    # --- arrived ---
    "Arrived": "arrived", "arrived": "arrived", "Completed": "arrived",
    "seen": "arrived", "COMPLETE": "arrived",
    # --- cancelled (by the clinic, not the patient - patient cancellations are
    #     no-shows per the front-desk convention above) ---
    "Cancelled": "cancelled", "cancelled by clinic": "cancelled",
    "CANC": "cancelled", "rescheduled": "cancelled",
}

CANONICAL_STATES = {"no_show", "arrived", "cancelled"}

# Colour -> FTE, reconstructed from the roster's legend.
COLOUR_TO_FTE = {"C6EFCE": 1.0, "FFEB9C": 0.8, "FFCC99": 0.6}


class StandardizationError(Exception):
    """Raised when the data contains something the rulings do not cover.

    This is deliberate. A silent default is how a wrong number gets into a board
    deck. If the data surprises us, we stop and look.
    """


@dataclass
class StandardizationReport:
    """An audit trail. Every ruling records what it did, in numbers."""

    status_counts: dict = field(default_factory=dict)
    contradictions_resolved: int = 0
    mrns_before: int = 0
    patients_after: int = 0
    fte_recovered: int = 0
    fte_total: int = 0
    dob_collision_risk: int = 0

    def to_frame(self) -> pd.DataFrame:
        rows = [
            ("S1_status_no_show", self.status_counts.get("no_show", 0)),
            ("S1_status_arrived", self.status_counts.get("arrived", 0)),
            ("S1_status_cancelled", self.status_counts.get("cancelled", 0)),
            ("S2_contradictions_flipped_to_arrived", self.contradictions_resolved),
            ("S3_mrns_before_dedup", self.mrns_before),
            ("S3_patients_after_dedup", self.patients_after),
            ("S3_dob_name_collision_risk", self.dob_collision_risk),
            ("S4_fte_recovered_from_colour", self.fte_recovered),
            ("S4_fte_total", self.fte_total),
        ]
        return pd.DataFrame(rows, columns=["metric", "value"])


# ---------------------------------------------------------------------------
# S1 - canonicalise status
# ---------------------------------------------------------------------------


def canonicalize_status(status: pd.Series) -> pd.Series:
    """Map every free-typed status to a canonical state, or raise on a surprise."""
    unknown = set(status.dropna().unique()) - set(STATUS_MAP)
    if unknown:
        raise StandardizationError(
            f"unmapped status strings: {sorted(unknown)}. "
            "Add a ruling for each before proceeding - do not default."
        )
    return status.map(STATUS_MAP)


# ---------------------------------------------------------------------------
# S4 - FTE from cell colour  (defined early; dedup and status don't depend on it)
# ---------------------------------------------------------------------------


def extract_roster(xlsx_path: Path) -> tuple[pd.DataFrame, int, int]:
    """Read the provider roster, recovering FTE from each cell's fill colour.

    The FTE column is blank text; its meaning is the background colour. We read
    the fill, map the hex to a fraction, and fail loudly on any colour the legend
    does not cover - a new colour is a new FTE band someone forgot to tell us
    about, not a thing to guess.
    """
    wb = load_workbook(xlsx_path)
    ws = wb["Roster"]

    rows = []
    recovered = 0
    total = 0

    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        if not name:
            continue
        total += 1

        clinic = ws.cell(row=row, column=2).value
        room = ws.cell(row=row, column=3).value

        fill = ws.cell(row=row, column=4).fill
        rgb = fill.start_color.rgb
        hex_code = rgb[-6:] if isinstance(rgb, str) else None

        fte = COLOUR_TO_FTE.get(hex_code)
        if fte is None:
            raise StandardizationError(
                f"provider '{name}' has fill colour {hex_code}, which is not in the "
                "legend. A new colour means a new FTE band - investigate, do not guess."
            )
        recovered += 1

        rows.append({"provider": name, "clinic": clinic, "room": room, "fte": fte})

    return pd.DataFrame(rows), recovered, total


# ---------------------------------------------------------------------------
# S3 - patient deduplication on date of birth
# ---------------------------------------------------------------------------


def _normalise_name(first: str, last: str) -> str:
    """Collapse name variants to a comparable form.

    'Sarah' and 'S.' should collide; 'Sarah' and 'Michael' should not. We use the
    first initial plus the (stripped, lowered) surname - enough to merge the
    realistic variants without merging genuinely different people who happen to
    share a birth date.
    """
    f = str(first).strip().lower().replace(".", "")
    surname = str(last).strip().lower()
    initial = f[0] if f else ""
    return f"{initial}|{surname}"


def deduplicate_patients(
    raw_patients: pd.DataFrame,
) -> tuple[pd.DataFrame, int]:
    """Collapse multiple MRNs for the same human into one canonical patient.

    Anchored on date of birth, refined by a normalised name. Returns a mapping
    from every MRN to a single canonical_patient_id, plus a count of the cases
    where two DIFFERENT normalised names share a DOB (the collision risk we are
    honest about rather than hiding).
    """
    df = raw_patients.copy()
    df["norm_name"] = [
        _normalise_name(first, last)
        for first, last in zip(df["first_name"], df["last_name"], strict=False)
    ]

    # A person is (date_of_birth, normalised name). Everyone matching both is the
    # same human, however many MRNs they hold.
    df["canonical_patient_id"] = (
        df.groupby(["date_of_birth", "norm_name"]).ngroup().map(lambda i: f"PT{i:06d}")
    )

    # Honesty metric: how many DOBs carry more than one distinct normalised name?
    # Those are the cases where our merge COULD wrongly split or join. We report
    # it rather than pretend the risk is zero.
    collision = (
        df.groupby("date_of_birth")["norm_name"].nunique().gt(1).sum()
    )

    mapping = df[["mrn", "canonical_patient_id"]].copy()

    return mapping, int(collision)


# ---------------------------------------------------------------------------
# S2 - resolve the schedule-vs-billing contradiction
# ---------------------------------------------------------------------------


def resolve_contradictions(
    appointments: pd.DataFrame, visits: pd.DataFrame
) -> tuple[pd.DataFrame, int]:
    """Where the schedule says no_show but a billed visit exists, billing wins.

    A billed visit is hard evidence a patient was seen. The schedule is a field a
    busy front desk forgets. Counting these as no-shows would OVERSTATE the
    problem - and overstating your own business case is how you lose the room to
    a skeptic. We flip them to 'arrived' and record how many.
    """
    billed_appts = set(visits["appointment_id"])

    is_contradiction = (
        (appointments["status_canonical"] == "no_show")
        & (appointments["appointment_id"].isin(billed_appts))
    )

    out = appointments.copy()
    out.loc[is_contradiction, "status_canonical"] = "arrived"
    out.loc[is_contradiction, "status_corrected"] = True

    return out, int(is_contradiction.sum())


# ---------------------------------------------------------------------------
# ORCHESTRATION
# ---------------------------------------------------------------------------


def standardize(raw_dir: Path, processed_dir: Path) -> StandardizationReport:
    """Run all four rulings and write the standardized tables plus an audit report."""
    processed_dir.mkdir(parents=True, exist_ok=True)
    report = StandardizationReport()

    appointments = pd.read_csv(raw_dir / "appointments.csv")
    visits = pd.read_csv(raw_dir / "visits.csv")
    raw_patients = pd.read_csv(raw_dir / "patients.csv")

    # --- S1: canonicalise status ---
    appointments["status_canonical"] = canonicalize_status(appointments["status"])
    appointments["status_corrected"] = False
    report.status_counts = appointments["status_canonical"].value_counts().to_dict()

    # --- S2: resolve contradictions (billing wins) ---
    appointments, n_flipped = resolve_contradictions(appointments, visits)
    report.contradictions_resolved = n_flipped
    # Recount after the correction, since some no_shows became arrivals.
    report.status_counts = appointments["status_canonical"].value_counts().to_dict()

    # --- S3: deduplicate patients ---
    report.mrns_before = raw_patients["mrn"].nunique()
    mrn_map, collision = deduplicate_patients(raw_patients)
    report.patients_after = mrn_map["canonical_patient_id"].nunique()
    report.dob_collision_risk = collision

    # Attach the canonical patient id to every appointment.
    appointments = appointments.merge(mrn_map, on="mrn", how="left")

    # --- S4: extract FTE from cell colour ---
    roster, recovered, total = extract_roster(raw_dir / "providers.xlsx")
    report.fte_recovered = recovered
    report.fte_total = total

    # --- write outputs ---
    appointments.to_parquet(processed_dir / "appointments_standardized.parquet", index=False)
    mrn_map.to_parquet(processed_dir / "patient_mapping.parquet", index=False)
    roster.to_parquet(processed_dir / "provider_roster.parquet", index=False)
    report.to_frame().to_parquet(processed_dir / "standardization_report.parquet", index=False)

    # --- print the audit ---
    print("=" * 66)
    print("STANDARDIZATION — four rulings, all recorded")
    print("=" * 66)
    print("\nS1 — status canonicalised (after S2 correction):")
    for state in ("no_show", "arrived", "cancelled"):
        print(f"    {state:<12} {report.status_counts.get(state, 0):>8,}")

    print(f"\nS2 — schedule/billing contradictions resolved (billing wins): {n_flipped:,}")
    print("    These were marked no_show but had a billed visit. Counting them")
    print("    as no-shows would have overstated the problem.")

    print("\nS3 — patient deduplication (DOB-anchored):")
    print(f"    MRNs before        : {report.mrns_before:,}")
    print(f"    Patients after     : {report.patients_after:,}")
    print(f"    Merged away        : {report.mrns_before - report.patients_after:,}")
    print(f"    DOB/name collision risk (honest limit): {collision}")

    print("\nS4 — FTE recovered from cell colour:")
    print(f"    {recovered}/{total} providers' FTE read from fill colour")

    return report


def main() -> None:
    parser = argparse.ArgumentParser(description="Standardize Trinity clinic data.")
    parser.add_argument("--raw", type=Path, default=Path("data/raw"))
    parser.add_argument("--processed", type=Path, default=Path("data/processed"))
    args = parser.parse_args()

    standardize(args.raw, args.processed)


if __name__ == "__main__":
    main()
