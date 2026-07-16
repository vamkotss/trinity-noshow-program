"""Seeded data generator for Trinity Family Health.

WHAT THIS BUILDS
----------------
The four operational data sources a real primary-care network would hand a
business analyst, each broken in the specific ways Tom (IT) warned about in
discovery:

  appointments.csv    scheduling export - 14 spellings of "no-show", statuses
                      that contradict the visit file, free-typed appt types
  visits.csv          billing extract - visits never marked arrived, typo'd
                      procedure codes, a payer column mixing names and plan IDs
  providers.xlsx      the artisanal roster - FTE encoded as CELL COLOUR, not data
  reminder_log.csv    free-text call outcomes, covering only 3 of the 7 clinics

WHY A GENERATOR AND NOT A DOWNLOAD
----------------------------------
Same reason as P1. The defects are known because we injected them, so the
standardization layer in M3 can be TESTED against a manifest rather than
eyeballed. "I cleaned the data" is a claim; "I recovered 14 status spellings
into 3 canonical states and here is the test that proves it" is evidence.

THE PLANTED BUSINESS TRUTH
--------------------------
No-shows are not random. Four drivers are built in, for the analyst to find:

  1. LEAD TIME. The longer between booking and appointment, the higher the
     no-show rate. This is the single strongest driver and the one that most
     justifies an overbooking policy.
  2. NEW PATIENTS. First-ever visits no-show far more than established patients.
  3. MONDAY MORNINGS. A weekday-and-time effect.
  4. PAYER. Self-pay and Medicaid no-show more than commercial - a real and
     ethically loaded pattern the analyst must handle carefully.

THE NATURAL EXPERIMENT (and its trap)
-------------------------------------
Three clinics logged reminder calls; four did not. Naively comparing no-show
rates between "reminded" and "not reminded" clinics looks like an experiment.
It is not - the three clinics that bothered to log reminders are also the
better-run clinics, so they would have had lower no-shows anyway. That
confound is planted deliberately. Detecting it is M5.

SYNTHETIC PHI
-------------
The raw data contains fake names, dates of birth, and phone numbers - exactly
what a real scheduling export leaks. M2 includes a de-identification pass; this
module generates the PHI so there is something realistic to de-identify.

Run:  python -m trinity.generate
"""

from __future__ import annotations

import argparse
from datetime import date, datetime, timedelta
from pathlib import Path

import numpy as np
import pandas as pd

# ---------------------------------------------------------------------------
# CONFIGURATION
# ---------------------------------------------------------------------------

SEED = 20260714
N_PATIENTS = 8_000
N_APPOINTMENTS = 60_000

START_DATE = date(2024, 7, 1)
END_DATE = date(2026, 6, 30)

# The seven clinics. Only the first three logged reminder calls - and, planted
# deliberately, they are also the better-run clinics with lower baseline
# no-show rates. That confound is the trap in the natural experiment.
CLINICS = {
    "Plano": {"reminds": True, "quality": 0.90},
    "Frisco": {"reminds": True, "quality": 0.92},
    "Allen": {"reminds": True, "quality": 0.88},
    "McKinney": {"reminds": False, "quality": 0.78},
    "Richardson": {"reminds": False, "quality": 0.80},
    "Garland": {"reminds": False, "quality": 0.76},
    "Denton": {"reminds": False, "quality": 0.79},
}

PROVIDERS = [
    ("Dr. Raghavan", "Plano", 1.0),
    ("Dr. Patel", "Plano", 1.0),
    ("Dr. Chen", "Frisco", 0.8),
    ("Dr. Okonkwo", "Frisco", 1.0),
    ("Dr. Martinez", "Allen", 0.6),
    ("Dr. Williams", "Allen", 1.0),
    ("Dr. Nguyen", "McKinney", 1.0),
    ("Dr. Foster", "McKinney", 0.8),
    ("Dr. Kim", "Richardson", 1.0),
    ("Dr. Brooks", "Garland", 0.6),
    ("Dr. Silva", "Garland", 1.0),
    ("Dr. Ahmed", "Denton", 0.8),
]

# FTE fraction -> the cell colour Brenda uses in the roster. This is the mapping
# the analyst must REVERSE by reading pixel colours out of the spreadsheet.
FTE_COLOURS = {
    1.0: "C6EFCE",   # green  = full-time
    0.8: "FFEB9C",   # yellow = part-time (0.8)
    0.6: "FFCC99",   # orange = per-diem (0.6)
}

APPT_TYPES_CLEAN = ["Annual Physical", "Follow-up", "Sick Visit", "New Patient", "Telehealth"]

# The same appointment type, free-typed the way a hurried front desk actually
# writes it. The analyst must collapse these.
APPT_TYPE_VARIANTS = {
    "Annual Physical": ["Annual Physical", "annual physical", "Physical", "AWV", "wellness"],
    "Follow-up": ["Follow-up", "followup", "F/U", "follow up", "FU"],
    "Sick Visit": ["Sick Visit", "sick", "acute", "Sick", "urgent"],
    "New Patient": ["New Patient", "new pt", "NP", "new patient", "NewPt"],
    "Telehealth": ["Telehealth", "telehealth", "TH", "video", "virtual"],
}

# THE FOURTEEN. Every string a human has ever typed to mean the patient did not
# show up. Recovering these into one canonical "no_show" state is the core of M3.
NO_SHOW_SPELLINGS = [
    "NS", "no show", "noshow", "no-show", "No Show", "DNA", "dna",
    "cancelled by pt", "CXL", "cxl", "did not attend", "n/s", "No-Show", "NOSHOW",
]

# Clean terminal states.
ARRIVED_SPELLINGS = ["Arrived", "arrived", "Completed", "seen", "COMPLETE"]
CANCELLED_SPELLINGS = ["Cancelled", "cancelled by clinic", "CANC", "rescheduled"]

PAYERS_CLEAN = ["BlueCross", "Aetna", "UnitedHealth", "Medicaid", "Medicare", "Self-Pay"]
# The payer column mixes payer NAMES with plan IDs - as a real export does when
# two systems feed the same field.
PAYER_PLAN_IDS = {"BlueCross": "BCBS-4410", "Aetna": "AET-0092", "UnitedHealth": "UHC-7731"}

# Procedure codes, and the typo'd versions that slip into a billing extract.
PROCEDURE_CODES = ["99213", "99214", "99215", "99203", "99204"]

DEFECT_RATES = {
    "duplicate_mrn": 0.04,           # 4% of patients get a second/third MRN
    "status_contradiction": 0.02,    # 2% of no-shows have a billed visit
    "typo_procedure_code": 0.05,     # 5% of procedure codes are malformed
    "payer_as_plan_id": 0.15,        # 15% of payers appear as a plan ID
    "missing_phone": 0.06,
}


def _rng(seed: int) -> np.random.Generator:
    return np.random.default_rng(seed)


# ---------------------------------------------------------------------------
# PATIENTS  (with synthetic PHI, to be de-identified in M2)
# ---------------------------------------------------------------------------


def build_patients(rng: np.random.Generator) -> pd.DataFrame:
    """One row per patient. Carries fake PHI on purpose."""
    first = ["James", "Mary", "Robert", "Patricia", "John", "Jennifer", "Michael",
             "Linda", "David", "Sarah", "Priya", "Wei", "Ana", "Omar", "Grace"]
    last = ["Smith", "Johnson", "Williams", "Brown", "Jones", "Garcia", "Miller",
            "Davis", "Rodriguez", "Martinez", "Chen", "Patel", "Okafor", "Nguyen"]

    n = N_PATIENTS
    fnames = rng.choice(first, n)
    lnames = rng.choice(last, n)

    # Date of birth - drives an "age" feature and is textbook PHI.
    ages = rng.integers(1, 90, n)
    dobs = [date(2026 - int(a), int(rng.integers(1, 13)), int(rng.integers(1, 29))) for a in ages]

    phones = [f"972-555-{rng.integers(1000, 9999)}" for _ in range(n)]
    # Some phones missing - the reminder log will struggle with these.
    missing = rng.random(n) < DEFECT_RATES["missing_phone"]
    phones = [None if m else p for p, m in zip(phones, missing, strict=False)]

    return pd.DataFrame(
        {
            "mrn": [f"MRN{i:06d}" for i in range(1, n + 1)],
            "first_name": fnames,
            "last_name": lnames,
            "date_of_birth": dobs,
            "phone": phones,
            "home_clinic": rng.choice(list(CLINICS), n),
            # A hidden "reliability" trait that drives this patient's no-show
            # propensity. NOT exported - the analyst must infer behaviour from
            # history, exactly as in real life.
            "reliability": rng.beta(6, 2, n),
        }
    )


# ---------------------------------------------------------------------------
# APPOINTMENTS  (the scheduling export - the messy heart of the project)
# ---------------------------------------------------------------------------


def _no_show_probability(
    lead_days: int, is_new: bool, weekday: int, hour: int, payer: str,
    reliability: float, clinic_quality: float,
) -> float:
    """The planted no-show model. The analyst reverse-engineers this from data.

    Every term here is a driver the analysis is meant to surface. The point is
    not that the model is fancy - it is that the drivers are REAL and RANKED, so
    "lead time is the biggest lever" is a finding, not a guess.
    """
    p = 0.08  # base rate

    # 1. LEAD TIME - the strongest driver. Booked far ahead => forgotten.
    p += min(lead_days, 60) * 0.0035

    # 2. NEW PATIENTS - never been, less committed.
    if is_new:
        p += 0.10

    # 3. MONDAY MORNINGS.
    if weekday == 0 and hour < 11:
        p += 0.06

    # 4. PAYER - a real, ethically loaded pattern.
    if payer in ("Self-Pay", "Medicaid"):
        p += 0.07

    # Patient's own reliability. The clinic-quality term is kept SMALL on
    # purpose: most of the reminding-clinics-look-better effect must flow through
    # the observable payer mix (assigned in build_appointments), so that an
    # analyst who stratifies on payer can actually recover the truth. If the
    # confound lived entirely in this hidden quality term, no amount of
    # adjustment on observable features could remove it, and the exercise would
    # teach the wrong lesson (that adjustment is futile).
    p -= (reliability - 0.7) * 0.25
    p -= (clinic_quality - 0.80) * 0.08

    return float(np.clip(p, 0.01, 0.85))


def build_appointments(
    patients: pd.DataFrame, rng: np.random.Generator
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """The scheduling export, and a clean shadow copy for the defect manifest."""
    prov_by_clinic: dict[str, list[str]] = {}
    for name, clinic, _fte in PROVIDERS:
        prov_by_clinic.setdefault(clinic, []).append(name)

    pat_records = patients.to_dict("records")
    seen_before: set[str] = set()

    rows = []
    for i in range(N_APPOINTMENTS):
        pat = pat_records[int(rng.integers(0, len(pat_records)))]
        clinic = pat["home_clinic"]

        booked = START_DATE + timedelta(days=int(rng.integers(0, (END_DATE - START_DATE).days - 60)))
        lead_days = int(rng.gamma(2.2, 6.0))
        appt_date = booked + timedelta(days=lead_days)
        if appt_date > END_DATE:
            continue

        hour = int(rng.integers(8, 17))
        weekday = appt_date.weekday()

        is_new = pat["mrn"] not in seen_before
        seen_before.add(pat["mrn"])

        # THE CONFOUND, made OBSERVABLE. The reminding clinics (the better-run,
        # more affluent-area ones) serve a lower-risk payer mix: fewer self-pay
        # and Medicaid patients. This is what actually drives their lower
        # no-show rate - not the reminding. Because payer is a column the analyst
        # CAN see, stratifying on it correctly shrinks the naive reminder effect
        # toward zero. That is the whole lesson: the confound is adjustable if you
        # look for it, and catastrophic if you do not.
        if CLINICS[clinic]["reminds"]:
            # Lower-risk mix: more commercial, less self-pay/Medicaid.
            payer_p = [0.34, 0.22, 0.22, 0.08, 0.11, 0.03]
        else:
            # Higher-risk mix.
            payer_p = [0.22, 0.14, 0.14, 0.20, 0.16, 0.14]
        payer = str(rng.choice(PAYERS_CLEAN, p=payer_p))

        p_ns = _no_show_probability(
            lead_days, is_new, weekday, hour, payer,
            pat["reliability"], CLINICS[clinic]["quality"],
        )

        roll = rng.random()
        if roll < p_ns:
            true_status = "no_show"
        elif roll < p_ns + 0.06:
            true_status = "cancelled"
        else:
            true_status = "arrived"

        appt_type_clean = str(rng.choice(APPT_TYPES_CLEAN))
        if is_new:
            appt_type_clean = "New Patient"

        rows.append(
            {
                "appointment_id": f"APT{i:07d}",
                "mrn": pat["mrn"],
                "clinic": clinic,
                "provider": str(rng.choice(prov_by_clinic[clinic])),
                "booked_date": booked,
                "appointment_date": appt_date,
                "appointment_hour": hour,
                "appointment_type_clean": appt_type_clean,
                "payer_clean": payer,
                "is_new_patient": is_new,
                "lead_days": lead_days,
                "true_status": true_status,
            }
        )

    clean = pd.DataFrame(rows)

    # ----- Now dirty it, the way the real export is dirty -----
    export = clean.copy()

    # Free-type the appointment type.
    export["appointment_type"] = [
        str(rng.choice(APPT_TYPE_VARIANTS[t])) for t in export["appointment_type_clean"]
    ]

    # Free-type the status. no_show gets one of the fourteen spellings.
    def dirty_status(s: str) -> str:
        if s == "no_show":
            return str(rng.choice(NO_SHOW_SPELLINGS))
        if s == "arrived":
            return str(rng.choice(ARRIVED_SPELLINGS))
        return str(rng.choice(CANCELLED_SPELLINGS))

    export["status"] = [dirty_status(s) for s in export["true_status"]]

    # Mix payer names with plan IDs.
    def dirty_payer(p: str) -> str:
        if p in PAYER_PLAN_IDS and rng.random() < DEFECT_RATES["payer_as_plan_id"]:
            return PAYER_PLAN_IDS[p]
        return p

    export["payer"] = [dirty_payer(p) for p in export["payer_clean"]]

    # Drop the clean helper columns from the export - the analyst does not get
    # the answer key. They stay in `clean` for the manifest and tests.
    export = export.drop(
        columns=["appointment_type_clean", "payer_clean", "true_status", "lead_days", "is_new_patient"]
    )

    return export, clean


# ---------------------------------------------------------------------------
# VISITS  (billing extract - the file that contradicts the schedule)
# ---------------------------------------------------------------------------


def build_visits(
    appointments_clean: pd.DataFrame, rng: np.random.Generator
) -> pd.DataFrame:
    """Billing rows for visits that happened - plus the contradiction defect.

    A billed visit exists for every 'arrived' appointment. THE DEFECT: a small
    fraction of appointments the schedule marks as no-show ALSO have a billed
    visit here. Someone arrived; the front desk never updated the schedule. The
    analyst must decide which source wins (ruling in M3).
    """
    arrived = appointments_clean[appointments_clean["true_status"] == "arrived"]

    # Some no-shows that were really arrivals - the contradiction.
    no_shows = appointments_clean[appointments_clean["true_status"] == "no_show"]
    n_contra = int(len(no_shows) * DEFECT_RATES["status_contradiction"])
    contradictions = no_shows.sample(n=n_contra, random_state=int(rng.integers(0, 10**6)))

    billed = pd.concat([arrived, contradictions], ignore_index=True)

    def dirty_code(_: object) -> str:
        code = str(rng.choice(PROCEDURE_CODES))
        # 5% get a typo: transposed digit, letter, or truncation.
        if rng.random() < DEFECT_RATES["typo_procedure_code"]:
            kind = rng.integers(0, 3)
            if kind == 0:
                code = code[:-1] + "O"          # letter O for zero
            elif kind == 1:
                code = code[1:] + code[0]        # rotate
            else:
                code = code[:4]                  # truncated
        return code

    return pd.DataFrame(
        {
            "visit_id": [f"VIS{i:07d}" for i in range(len(billed))],
            "appointment_id": billed["appointment_id"].to_numpy(),
            "mrn": billed["mrn"].to_numpy(),
            "visit_date": billed["appointment_date"].to_numpy(),
            "procedure_code": [dirty_code(None) for _ in range(len(billed))],
            "billed_amount": np.round(rng.uniform(80, 320, len(billed)), 2),
        }
    )


# ---------------------------------------------------------------------------
# REMINDER LOG  (3 of 7 clinics - the natural experiment)
# ---------------------------------------------------------------------------


def build_reminder_log(
    appointments_clean: pd.DataFrame, rng: np.random.Generator
) -> pd.DataFrame:
    """Free-text reminder-call outcomes, for the three clinics that logged them.

    The four clinics without a log are simply absent - which is exactly the
    selection bias. The clinics that logged reminders are the better-run ones.
    """
    reminding = [c for c, v in CLINICS.items() if v["reminds"]]
    subset = appointments_clean[appointments_clean["clinic"].isin(reminding)]

    # Not every appointment got a call - "when there's time", per Marcus.
    called = subset.sample(frac=0.7, random_state=int(rng.integers(0, 10**6)))

    outcomes_free = {
        "confirmed": ["confirmed", "Confirmed", "conf", "pt confirmed", "C"],
        "voicemail": ["lvm", "left voicemail", "LVM", "voicemail", "no answer left msg"],
        "no_answer": ["no answer", "NA", "n/a", "didnt pick up", "no ans"],
        "wrong_number": ["wrong #", "wrong number", "WN", "bad number", "disconnected"],
    }

    def outcome(_: object) -> str:
        kind = str(rng.choice(list(outcomes_free), p=[0.55, 0.25, 0.15, 0.05]))
        return str(rng.choice(outcomes_free[kind]))

    return pd.DataFrame(
        {
            "call_id": [f"CALL{i:07d}" for i in range(len(called))],
            "appointment_id": called["appointment_id"].to_numpy(),
            "clinic": called["clinic"].to_numpy(),
            "call_outcome": [outcome(None) for _ in range(len(called))],
            "call_date": [
                d - timedelta(days=int(rng.integers(1, 3)))
                for d in called["appointment_date"]
            ],
        }
    )


# ---------------------------------------------------------------------------
# DUPLICATE MRNs  (same human, several record numbers)
# ---------------------------------------------------------------------------


def inject_duplicate_mrns(
    patients: pd.DataFrame, appointments: pd.DataFrame, rng: np.random.Generator
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Give some patients a second/third MRN with a name variant.

    Then rewrite a share of their appointments to use the alias MRN, so the same
    human appears under multiple record numbers - and per-patient no-show rates
    are wrong until the analyst dedupes them.
    """
    n_dupes = int(len(patients) * DEFECT_RATES["duplicate_mrn"])
    dupe_src = patients.sample(n=n_dupes, random_state=int(rng.integers(0, 10**6)))

    alias_rows = []
    mapping_rows = []
    next_id = len(patients) + 1

    for pat in dupe_src.itertuples(index=False):
        alias_mrn = f"MRN{next_id:06d}"
        next_id += 1

        # Same person, slightly different name - the realistic dedupe challenge.
        fn = pat.first_name
        if rng.random() < 0.5 and len(fn) > 3:
            fn = fn[0] + "."          # "Sarah" -> "S."
        ln = pat.last_name
        if rng.random() < 0.3:
            ln = ln + " "             # trailing space

        alias_rows.append(
            {
                "mrn": alias_mrn,
                "first_name": fn,
                "last_name": ln,
                "date_of_birth": pat.date_of_birth,   # same DOB - the join key
                "phone": pat.phone,
                "home_clinic": pat.home_clinic,
                "reliability": pat.reliability,
            }
        )
        mapping_rows.append({"alias_mrn": alias_mrn, "true_mrn": pat.mrn})

        # Move ~40% of this patient's appointments onto the alias.
        their_appts = appointments[appointments["mrn"] == pat.mrn]
        move = their_appts.sample(frac=0.4, random_state=int(rng.integers(0, 10**6)))
        appointments.loc[move.index, "mrn"] = alias_mrn

    patients_out = pd.concat([patients, pd.DataFrame(alias_rows)], ignore_index=True)
    mapping = pd.DataFrame(mapping_rows)

    return patients_out, appointments, mapping


# ---------------------------------------------------------------------------
# THE EXCEL ROSTER  (FTE as cell colour - the horrible realistic task)
# ---------------------------------------------------------------------------


def write_provider_roster(path: Path) -> pd.DataFrame:
    """Write a real .xlsx where each provider's FTE is a CELL BACKGROUND COLOUR.

    There is no FTE column. Green means 1.0, yellow 0.8, orange 0.6. The analyst
    must read the fill colour out of each cell with openpyxl and map it back to a
    number. This is not a contrived exercise - operational data lives in
    spreadsheets like this constantly, and "I extracted structured data from cell
    formatting" is a genuinely differentiating line.

    Returns the ground-truth FTE table for the test manifest.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Roster"

    headers = ["Provider", "Clinic", "Room", "FTE (see colour key)"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    truth = []
    for idx, (name, clinic, fte) in enumerate(PROVIDERS, start=2):
        room = f"Rm {100 + idx}"
        ws.cell(row=idx, column=1, value=name)
        ws.cell(row=idx, column=2, value=clinic)
        ws.cell(row=idx, column=3, value=room)

        # THE DEFECT: the FTE cell is blank. Its meaning is the FILL COLOUR.
        fte_cell = ws.cell(row=idx, column=4, value="")
        colour = FTE_COLOURS[fte]
        fte_cell.fill = PatternFill(start_color=colour, end_color=colour, fill_type="solid")

        truth.append({"provider": name, "clinic": clinic, "room": room, "fte": fte})

    # A legend, off to the side, as Brenda would leave it - human-readable only.
    ws.cell(row=2, column=6, value="Colour key:")
    for i, (fte, colour) in enumerate(FTE_COLOURS.items(), start=3):
        key = ws.cell(row=i, column=6, value=f"= {fte} FTE")
        key.fill = PatternFill(start_color=colour, end_color=colour, fill_type="solid")

    wb.save(path)
    return pd.DataFrame(truth)


# ---------------------------------------------------------------------------
# ORCHESTRATION
# ---------------------------------------------------------------------------


def generate(out_dir: Path, seed: int = SEED) -> dict[str, int]:
    rng = _rng(seed)
    out_dir.mkdir(parents=True, exist_ok=True)

    print(f"Seed: {seed}")
    print("Building patients (with synthetic PHI)...")
    patients = build_patients(rng)

    print("Building appointments (the messy scheduling export)...")
    appointments, appts_clean = build_appointments(patients, rng)

    print("Injecting duplicate MRNs...")
    patients, appointments, mrn_map = inject_duplicate_mrns(patients, appointments, rng)

    print("Building visits (billing extract, with contradictions)...")
    visits = build_visits(appts_clean, rng)

    print("Building reminder log (3 of 7 clinics)...")
    reminders = build_reminder_log(appts_clean, rng)

    print("Writing provider roster (FTE as cell colour)...")
    roster_truth = write_provider_roster(out_dir / "providers.xlsx")

    # The manifest: ground truth the standardization layer is tested against.
    # It stays OUT of the files an analyst would be handed - it is the answer key.
    manifest = {
        "n_no_show_spellings": len(NO_SHOW_SPELLINGS),
        "n_duplicate_mrns": len(mrn_map),
        "reminding_clinics": [c for c, v in CLINICS.items() if v["reminds"]],
    }

    # Patients export drops the hidden reliability trait.
    patients_export = patients.drop(columns=["reliability"])

    files = {
        "patients.csv": patients_export,
        "appointments.csv": appointments,
        "visits.csv": visits,
        "reminder_log.csv": reminders,
        "_mrn_truth.csv": mrn_map,           # answer key, prefixed with _
        "_roster_truth.csv": roster_truth,   # answer key
    }

    counts = {}
    for name, df in files.items():
        df.to_csv(out_dir / name, index=False)
        counts[name] = len(df)
        tag = "KEY " if name.startswith("_") else "DATA"
        print(f"  [{tag}] {name:24s} {len(df):>8,} rows")

    print(f"  [DATA] providers.xlsx           {len(roster_truth):>8,} rows  (FTE = cell colour)")

    print("\nManifest:")
    for k, v in manifest.items():
        print(f"  {k}: {v}")

    return counts


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate messy Trinity clinic data.")
    parser.add_argument("--out", type=Path, default=Path("data/raw"))
    parser.add_argument("--seed", type=int, default=SEED)
    args = parser.parse_args()

    started = datetime.now()
    generate(args.out, args.seed)
    print(f"\nDone in {(datetime.now() - started).total_seconds():.1f}s")


if __name__ == "__main__":
    main()
